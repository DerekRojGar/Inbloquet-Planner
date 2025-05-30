#formularios_view.py
import streamlit as st
from models.actividades_model import (
    cargar_escuelas, agregar_escuela,
    cargar_grupos, agregar_grupo,
    cargar_alumnos_activos, DIAS_ESPAÑOL
)
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

def render_new_activity_form():
    st.markdown("---")
    st.subheader("📝 Nueva Actividad")
    new_activity = None

    # 1. Inicializar estados esenciales
    if 'grupos_disponibles' not in st.session_state:
        st.session_state.grupos_disponibles = []
    if 'form_data' not in st.session_state:
        st.session_state.form_data = {
            'horario': '',
            'alumnos': '',
            'escuelas': [],
            'grupos': [],
            'encargado': '',
            'tema': '',
            'maestro': '',
            'notas': ''
        }

    escuelas = cargar_escuelas()
    grupos_por_escuela = cargar_grupos()
    alumnos_activos = cargar_alumnos_activos()

    # --- ESCUELAS: Agregar fuera del form ---
    nueva_escuela = st.text_input("Agregar nueva escuela:", "", key="nueva_escuela")
    if st.button("➕ Agregar Escuela"):
        if nueva_escuela and nueva_escuela not in escuelas:
            agregar_escuela(nueva_escuela)
            escuelas = cargar_escuelas()
            st.success(f"Escuela '{nueva_escuela}' agregada.")
            st.rerun()

    # --- GRUPOS: Agregar fuera del form ---
    escuela_para_grupo = st.selectbox("Selecciona escuela para agregar grupo:", escuelas, key="escuela_para_grupo")
    nuevo_grupo = st.text_input("Agregar nuevo grupo/taller:", "", key="nuevo_grupo")
    if st.button("➕ Agregar Grupo"):
        if nuevo_grupo:
            agregar_grupo(escuela_para_grupo, nuevo_grupo)
            st.success(f"Grupo '{nuevo_grupo}' agregado a '{escuela_para_grupo}'.")
            st.rerun()

    with st.form(key='form_nueva_actividad'):
        cols = st.columns(2)
        with cols[0]:
            dia_seleccionado = st.selectbox("Día:", DIAS_ESPAÑOL)
            dia_index = DIAS_ESPAÑOL.index(dia_seleccionado)

            horario = st.text_input("Horario:",
                                   value=st.session_state.form_data['horario'],
                                   placeholder="Ej: 8:00 a.m. - 2:00 p.m.")

            escuelas_seleccionadas = st.multiselect(
                "Escuelas:",
                escuelas,
                default=st.session_state.form_data['escuelas'],
                key="escuelas_multiselect"
            )

            # --- BUSCAR GRUPOS ---
            buscar_grupos = st.form_submit_button("🔍 Buscar grupos")
            if buscar_grupos:
                grupos_disponibles = []
                for esc in escuelas_seleccionadas:
                    grupos_disponibles.extend(grupos_por_escuela.get(esc, []))
                grupos_disponibles = sorted(set(grupos_disponibles))
                # Limpiar grupos seleccionados si ya no pertenecen
                grupos_seleccionados_validos = [
                    g for g in st.session_state.form_data['grupos'] if g in grupos_disponibles
                ]
                st.session_state.form_data['grupos'] = grupos_seleccionados_validos
                st.session_state.grupos_disponibles = grupos_disponibles
                st.rerun()
            else:
                grupos_disponibles = st.session_state.get('grupos_disponibles', [])
                if not grupos_disponibles:
                    for esc in escuelas_seleccionadas:
                        grupos_disponibles.extend(grupos_por_escuela.get(esc, []))
                    grupos_disponibles = sorted(set(grupos_disponibles))

            grupos_seleccionados = st.multiselect(
                "Grupos:",
                grupos_disponibles,
                default=st.session_state.form_data['grupos'],
                key="grupos_multiselect"
            )
            st.session_state.form_data['grupos'] = grupos_seleccionados

            # --- ALUMNOS ---
            # Si solo INBLOQUET está seleccionada, mostrar alumnos activos de INBLOQUET
            if "INBLOQUET" in escuelas_seleccionadas and len(escuelas_seleccionadas) == 1 and alumnos_activos:
                alumnos = st.multiselect(
                    "Alumnos (solo activos INBLOQUET):",
                    alumnos_activos,
                    default=st.session_state.form_data['alumnos'].split(", ") if st.session_state.form_data['alumnos'] else [],
                    key="alumnos_multiselect"
                )
                alumnos_str = ", ".join(alumnos)
            else:
                alumnos_str = st.text_input(
                    "Alumnos:",
                    value=st.session_state.form_data['alumnos'],
                    placeholder="Ej: Dereck, Alberto, Emma"
                )

        with cols[1]:
            encargado = st.text_input("Encargado:",
                                     value=st.session_state.form_data['encargado'],
                                     placeholder="Ej: Ivan, Gus, Caleb")

            tema = st.text_area("Descripción:",
                              value=st.session_state.form_data['tema'],
                              placeholder="Ej: Taller de programación con robots",
                              height=100)

            maestro = st.text_input("Maestro:",
                                  value=st.session_state.form_data['maestro'],
                                  placeholder="Ej: Joss, Angie, Kevin")

            notas = st.text_area("Notas:",
                               value=st.session_state.form_data['notas'],
                               placeholder="Ej: Materiales especiales requeridos")

        # 3. Botón de guardar con persistencia de datos
        if st.form_submit_button("💾 Guardar Actividad"):
            if escuelas_seleccionadas:
                new_activity = {
                    "Horario": horario or "-",
                    "Alumnos": alumnos_str or "-",
                    "Escuelas": ", ".join(escuelas_seleccionadas),
                    "Grupos": ", ".join(grupos_seleccionados) if grupos_seleccionados else "-",
                    "Maestro": maestro or "-",
                    "Tema": tema or "-",
                    "Encargado": encargado or "-",
                    "Notas": notas or "-",
                    "dia_index": dia_index
                }
                
                # Guardar en la estructura principal
                semana_key = f"{st.session_state.num_año}-S{st.session_state.num_semana}"
                fecha_key = st.session_state.actividades[semana_key]["fechas"][dia_index]
                st.session_state.actividades[semana_key]["actividades"][fecha_key].append(new_activity)
                
                # Persistir en CSV
                from models.actividades_model import guardar_datos
                guardar_datos(st.session_state.actividades)
                
                # Limpiar formulario
                st.session_state.form_data = {
                    'horario': '',
                    'alumnos': '',
                    'escuelas': [],
                    'grupos': [],
                    'encargado': '',
                    'tema': '',
                    'maestro': '',
                    'notas': ''
                }
                st.session_state.grupos_disponibles = []
                st.rerun()
            else:
                st.error("❌ Debes seleccionar al menos una escuela")

    return new_activity

def render_edit_activity_form():
    from models.actividades_model import guardar_datos, cargar_escuelas, cargar_grupos, cargar_alumnos_activos
    
    st.markdown("---")
    st.subheader("✏️ Editor de Actividad")
    ESCUELAS = cargar_escuelas()
    GRUPOS_POR_ESCUELA = cargar_grupos()
    ALUMNOS_ACTIVOS = cargar_alumnos_activos()
    
    if "editando" not in st.session_state:
        st.error("⚠️ No hay actividad seleccionada para editar")
        return
    
    edit = st.session_state.editando
    actividad = edit["datos"]
    semana_key = edit["semana_key"]
    fecha_original = edit["dia_str"]
    index_original = edit["index"]

    # 1. Obtener el índice del día original
    try:
        fecha_index = st.session_state.actividades[semana_key]["fechas"].index(fecha_original)
        dia_original = DIAS_ESPAÑOL[fecha_index]
    except ValueError:
        dia_original = DIAS_ESPAÑOL[0]
        fecha_index = 0

    with st.form(key='form_edicion_actividad'):
        cols = st.columns(2)
        
        with cols[0]:
            # 2. Selectbox con día actual correcto
            dia_seleccionado = st.selectbox(
                "Día:", 
                DIAS_ESPAÑOL, 
                index=fecha_index,
                key="edit_dia"
            )
            
            # 3. Obtener el nuevo índice del día seleccionado
            nuevo_fecha_index = DIAS_ESPAÑOL.index(dia_seleccionado)
            nueva_fecha = st.session_state.actividades[semana_key]["fechas"][nuevo_fecha_index]
            
            horario = st.text_input(
                "Horario:", 
                value=str(actividad.get("Horario", "-")).replace("-", "").strip(),
                key="edit_horario"
            )
            
            escuelas_actuales = [e.strip() for e in str(actividad.get("Escuelas", "")).split(",") if e.strip() in ESCUELAS]
            escuelas_edit = st.multiselect(
                "Escuelas:", 
                ESCUELAS, 
                default=escuelas_actuales,
                key="edit_escuelas"
            )
            
            if st.form_submit_button("🔄 Actualizar grupos"):
                grupos_disponibles = set()
                for escuela in escuelas_edit:
                    grupos_disponibles.update(GRUPOS_POR_ESCUELA.get(escuela, []))
                st.session_state.grupos_edit_disponibles = sorted(grupos_disponibles)
                st.rerun()
            
            grupos_disponibles = st.session_state.get("grupos_edit_disponibles", [])
            grupos_actuales = [g.strip() for g in str(actividad.get("Grupos", "")).split(",") if g.strip() in grupos_disponibles]
            grupos_edit = st.multiselect(
                "Grupos:", 
                grupos_disponibles,
                default=grupos_actuales,
                key="edit_grupos"
            )
            
            # --- ALUMNOS ---
            # Si solo INBLOQUET está seleccionada, mostrar alumnos activos de INBLOQUET
            if "INBLOQUET" in escuelas_edit and len(escuelas_edit) == 1 and ALUMNOS_ACTIVOS:
                alumnos_edit_list = [a.strip() for a in str(actividad.get("Alumnos", "-")).replace("-", "").strip().split(", ")]
                # Filtrar solo los que están en ALUMNOS_ACTIVOS para evitar error de Streamlit
                alumnos_edit_list = [a for a in alumnos_edit_list if a in ALUMNOS_ACTIVOS]
                alumnos_edit = st.multiselect(
                    "Alumnos (solo activos INBLOQUET):",
                    ALUMNOS_ACTIVOS,
                    default=alumnos_edit_list,
                    key="edit_alumnos_multiselect"
                )
                alumnos_edit_str = ", ".join(alumnos_edit)
            else:
                alumnos_edit_str = st.text_input(
                    "Alumnos:",
                    value=str(actividad.get("Alumnos", "-")).replace("-", "").strip(),
                    key="edit_alumnos"
                )

        with cols[1]:
            encargado_edit = st.text_input(
                "Encargado:", 
                value=str(actividad.get("Encargado", "-")).replace("-", "").strip(),
                key="edit_encargado"
            )
            maestro_edit = st.text_input(
                "Maestro:", 
                value=str(actividad.get("Maestro", "-")).replace("-", "").strip(),
                key="edit_maestro"
            )
            
            tema_edit = st.text_area(
                "Tema:", 
                value=str(actividad.get("Tema", "-")).replace("-", "").strip(),
                height=100,
                key="edit_tema"
            )
            
            notas_edit = st.text_area(
                "Notas:", 
                value=str(actividad.get("Notas", "-")).replace("-", "").strip(),
                height=100,
                key="edit_notas"
            )

        submit_cols = st.columns([1, 1, 3])
        with submit_cols[0]:
            submit = st.form_submit_button("💾 Guardar Cambios")
        with submit_cols[1]:
            cancelar = st.form_submit_button("❌ Cancelar")

        if submit:
            if escuelas_edit:
                nueva_actividad = {
                    "Día": dia_seleccionado,
                    "Horario": horario.strip() or "-",
                    "Escuelas": ", ".join(escuelas_edit),
                    "Grupos": ", ".join(grupos_edit) if grupos_edit else "-",
                    "Alumnos": alumnos_edit_str.strip() or "-",
                    "Encargado": encargado_edit.strip() or "-",
                    "Maestro": maestro_edit.strip() or "-",
                    "Tema": tema_edit.strip() or "-",
                    "Notas": notas_edit.strip() or "-"
                }
                
                # 4. Manejar cambio de día
                if nueva_fecha != fecha_original:
                    # Mover actividad a nueva fecha
                    actividad_eliminada = st.session_state.actividades[semana_key]["actividades"][fecha_original].pop(index_original)
                    st.session_state.actividades[semana_key]["actividades"][nueva_fecha].append(nueva_actividad)
                else:
                    # Actualizar en mismo día
                    st.session_state.actividades[semana_key]["actividades"][fecha_original][index_original] = nueva_actividad
                
                guardar_datos(st.session_state.actividades)
                
                # Limpiar estados
                for key in ['editando', 'selected_activity']:
                    if key in st.session_state:
                        del st.session_state[key]
                
                st.rerun()
            else:
                st.error("❌ Debes seleccionar al menos una escuela")

        if cancelar:
            del st.session_state.editando
            st.rerun()

def render_export_view(semana_key, semana, año):
    import base64
    import streamlit as st
    from models.export_model import crear_excel_con_diseño
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.drawing.image import Image
    from openpyxl import Workbook

    st.markdown("---")
    # Exportación de la semana actual (con diseño original)
    if st.button("📤 Exportar a Excel y CSV", key="export_excel_csv"):
        all_data = []
        semana_data = st.session_state.actividades[semana_key]
        for fecha in semana_data["fechas"]:
            dia_nombre = DIAS_ESPAÑOL[semana_data["fechas"].index(fecha)]
            for act in semana_data["actividades"][fecha]:
                registro = {
                    "Semana": semana,
                    "Año": año,
                    "Fecha": fecha,
                    "Día": dia_nombre,
                    **act
                }
                all_data.append(registro)
        df = pd.DataFrame(all_data)
        excel_file = f"Planificacion_S{semana}_{año}.xlsx"
        csv_file = f"Planificacion_S{semana}_{año}.csv"
        # Usamos la función de exportación con diseño para una semana
        crear_excel_con_diseño(df, excel_file, semana, año)
        df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        excel_base64 = base64.b64encode(open(excel_file, "rb").read()).decode()
        csv_base64 = base64.b64encode(open(csv_file, "rb").read()).decode()
        st.markdown("### 📥 Descargas")
        st.markdown(f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{excel_file}">Descargar Excel</a>', unsafe_allow_html=True)
        st.markdown(f'<a href="data:text/csv;base64,{csv_base64}" download="{csv_file}">Descargar CSV</a>', unsafe_allow_html=True)
        st.success("✅ Exportación completada")
    
    # Exportación de todas las semanas con el mismo diseño (sin Rocket)
    if st.button("📤 Exportar Semanas Totales", key="export_all_weeks"):
        wb = Workbook()
        # Eliminar hoja por defecto
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        semanas_ordenadas = sorted(
            st.session_state.actividades.keys(),
            key=lambda k: (int(k.split("-S")[0]), int(k.split("-S")[1]))
        )
        total_semanas = len(semanas_ordenadas)
        progreso = st.progress(0)
        
        # Recorremos cada semana y agregamos una hoja estilizada
        for idx, s_key in enumerate(semanas_ordenadas):
            año_str, semana_str = s_key.split("-S")
            anio_int = int(año_str)
            semana_int = int(semana_str)
            semana_data = st.session_state.actividades[s_key]
            all_data = []
            for fecha in semana_data["fechas"]:
                dia_nombre = DIAS_ESPAÑOL[semana_data["fechas"].index(fecha)]
                for act in semana_data["actividades"][fecha]:
                    registro = {
                        "Semana": semana_int,
                        "Año": anio_int,
                        "Fecha": fecha,
                        "Día": dia_nombre,
                        **act
                    }
                    all_data.append(registro)
            # Si hay datos para la semana, procesarlos
            if all_data:
                df = pd.DataFrame(all_data)
                # Creamos una nueva hoja para la semana en el Workbook
                ws = wb.create_sheet(title=f"S{semana_int}_{anio_int}")
                ws.insert_rows(1, 3)
                ws.merge_cells("A1:B3")
                ws.merge_cells("D1:I1")
                ws.merge_cells("D2:I2")
                ws.merge_cells("D3:I3")
                ws.merge_cells("A4:K4")
                
                # Insertar logo INBLOQUET (pero NO Rocket)
                try:
                    img = Image("InbloquetD.png")
                    img.width = 150
                    img.height = 140
                    ws.add_image(img, "A1")
                    row_height = img.height / 4
                except FileNotFoundError:
                    ws["A1"] = "LOGO NO ENCONTRADO"
                    ws["A1"].font = Font(bold=True, size=14, color="003891")
                    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    row_height = 20

                ws.row_dimensions[1].height = row_height
                ws.row_dimensions[2].height = row_height
                ws.row_dimensions[3].height = row_height

                # Encabezado con estilos (idéntico al de una semana)
                ws["D1"] = "Programación de Actividades Semanal"
                ws["D1"].font = Font(bold=True, size=20, color="003891", name="Gotham")
                ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

                ws["D2"] = f"SEMANA {semana_int} - AÑO {anio_int}"
                ws["D2"].font = Font(bold=True, size=16, color="4bb1e0", name="Gotham")
                ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

                ws["D3"] = "¡Intenta, Explora y Conquista!"
                ws["D3"].font = Font(bold=True, size=14, color="f6c500", name="Gotham")
                ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

                # Formatear datos
                df["Fecha"] = df.apply(lambda row: f"{row['Día']} {int(row['Fecha'][:2])} de {datetime.strptime(row['Fecha'], '%d/%m').strftime('%B').capitalize()} del {row['Año']}", axis=1)
                df = df.drop(columns=["Día", "Año"])
                column_order = ["Semana", "Fecha", "Horario", "Alumnos", "Escuelas", "Grupos", "Maestro", "Tema", "Encargado", "Notas"]
                df = df[column_order]
                ws.append(df.columns.tolist())
                
                # Aplicar color a encabezados
                header_fill = PatternFill(start_color="003891", end_color="003891", fill_type="solid")
                for cell in ws[5]:
                    cell.fill = header_fill
                    cell.font = Font(name="Gotham", bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Añadir datos
                for row in df.to_dict('records'):
                    ws.append(list(row.values()))
                
                # Crear tabla con estilo
                from openpyxl.worksheet.table import Table, TableStyleInfo
                last_row = ws.max_row
                tabla = Table(
                    displayName=f"TablaActividades_S{semana_int}_{anio_int}",
                    ref=f"A5:J{last_row}",
                    tableStyleInfo=TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                )
                ws.add_table(tabla)
                
                # Aplicar color a filas alternas
                for row_idx in range(6, ws.max_row + 1):
                    if row_idx % 2 == 0:
                        for cell in ws[row_idx]:
                            cell.fill = PatternFill(start_color="f6c500", end_color="f6c500", fill_type="solid")
                
                # Ajustar ancho de columnas
                for col in ws.iter_cols(min_row=5, max_row=ws.max_row):
                    if col[0].column_letter:
                        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = max_length + 2

            progreso.progress((idx + 1) / total_semanas)
        
        progreso.empty()
        excel_file = "Planificacion General de Semanas.xlsx"
        wb.save(excel_file)
        with open(excel_file, "rb") as f:
            excel_bytes = f.read()
        excel_base64 = base64.b64encode(excel_bytes).decode()
        st.markdown("### 📥 Descargas")
        st.markdown(f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{excel_file}">Descargar Excel</a>', unsafe_allow_html=True)
        st.success("✅ Exportación completada")

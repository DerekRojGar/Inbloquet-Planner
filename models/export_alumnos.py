import locale
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image

# Paleta de colores corporativos (la misma que en actividades)
COLORES_INBLOQUET = {
    "azul_principal": "003891",    # Azul fuerte
    "azul_secundario": "4BB1E0",   # Azul claro
    "amarillo": "F6C500",
    "verde": "D1DD00"
}


def generar_nombre_archivo_individual(alumno: dict) -> str:
    """
    Genera un nombre de archivo único para un alumno individual.
    Ejemplo: Ficha_Juan_Perez_123.xlsx
    """
    nombre = alumno.get("Nombre", "Alumno").strip().replace(" ", "_")
    matricula = alumno.get("Matricula", "0").strip()
    return f"Ficha_{nombre}_{matricula}.xlsx"

def generar_nombre_archivo_global() -> str:
    """
    Genera un nombre de archivo para exportar todos los alumnos.
    Ejemplo: Matricula_Completa_20250414.xlsx
    """
    fecha_str = datetime.now().strftime("%Y%m%d")
    return f"Matricula_Completa_{fecha_str}.xlsx"


def exportar_alumno_inbloquet(alumno: dict, filename: str):
    """
    Exporta UN alumno a Excel con el diseño y colores de Inbloquet, 
    replicando el formato de actividades: encabezado con logo y merges; 
    tres tablas para: Datos, Pagos y Clases; con filas alternas en amarillo.
    """
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    except locale.Error:
        locale.setlocale(locale.LC_TIME, "Spanish_Spain.1252")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "FichaAlumno"

    # --- Encabezado ---
    ws.insert_rows(1, 3)
    ws.merge_cells("A1:B3")   # Logo
    ws.merge_cells("D1:I1")
    ws.merge_cells("D2:I2")
    ws.merge_cells("D3:I3")
    ws.merge_cells("A4:K4")
    
    try:
        img = Image("InbloquetD.png")
        img.width = 120
        img.height = 120
        ws.add_image(img, "A1")
        row_height = img.height / 4
    except FileNotFoundError:
        ws["A1"] = "LOGO INBLOQUET"
        ws["A1"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["azul_principal"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row_height = 20
    for i in range(1, 4):
        ws.row_dimensions[i].height = row_height

    ws["D1"] = "Ficha Informativa del Alumno"
    ws["D1"].font = Font(bold=True, size=20, color=COLORES_INBLOQUET["azul_principal"])
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D2"] = f"ALUMNO: {alumno.get('Nombre','')}  -  Matrícula: {alumno.get('Matricula','')}"
    ws["D2"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["azul_secundario"])
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D3"] = "¡Intenta, Explora y Conquista!"
    ws["D3"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["amarillo"])
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Tabla de Datos del Alumno ---
    campos = [
        ("Grado", alumno.get("Grado","")),
        ("Grado Inbloquet", alumno.get("Grado Inbloquet","")),
        ("Curso", alumno.get("Tipo de Curso","")),
        ("Fecha Inscripción", alumno.get("Fecha Inscripción","")),
        ("Fecha Inicio Clases", alumno.get("Fecha Inicio Clases","")),
        ("Estado Inscripción", alumno.get("Inscripción","")),
        ("Escuela Procedencia", alumno.get("Escuela de provinencia","")),
        ("Alergias", alumno.get("Alergias","Ninguna")),
        ("Observaciones", alumno.get("Observaciones","Ninguna")),
        ("Contacto Familiar", alumno.get("Nombre Completo del Familiar","")),
        ("Tel. Familiar", alumno.get("Número de teléfono del familiar","")),
        ("Contacto Emergencia", alumno.get("Nombre completo de contacto de emergencia","")),
        ("Tel. Emergencia", alumno.get("Número de teléfono de contacto de emergencia",""))
    ]
    start_data = ws.max_row + 1
    # Encabezados de la tabla de datos
    ws.append(["Campo", "Valor"])
    for campo in campos:
        ws.append(list(campo))
    end_data = ws.max_row
    tabla_datos = Table(
        displayName=f"TablaDatos_{alumno.get('Matricula','')}",
        ref=f"A{start_data}:B{end_data}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws.add_table(tabla_datos)
    # Aplicar bandas amarillas a filas pares de la tabla de datos
    for row_idx in range(start_data + 1, end_data + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(
                    start_color=COLORES_INBLOQUET["amarillo"],
                    end_color=COLORES_INBLOQUET["amarillo"],
                    fill_type="solid"
                )

    # --- Tabla de Pagos ---
    pagos = alumno.get("pagos", [])
    df_pagos = pd.DataFrame(pagos, columns=["fecha", "monto", "metodo"])
    if df_pagos.empty:
        df_pagos = pd.DataFrame([["", "", ""]], columns=["fecha", "monto", "metodo"])

    ws.append([])  # Fila en blanco
    titulo_pagos_row = ws.max_row + 1
    ws.cell(row=titulo_pagos_row, column=1, value="Historial de Pagos").font = Font(bold=True, size=12, color=COLORES_INBLOQUET["azul_principal"])
    start_pagos = ws.max_row + 1
    ws.append(["Fecha", "Monto", "Método"])
    for r in df_pagos.itertuples(index=False):
        ws.append(list(r))
    end_pagos = ws.max_row
    tabla_pagos = Table(
        displayName=f"TablaPagos_{alumno.get('Matricula','')}",
        ref=f"A{start_pagos}:C{end_pagos}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws.add_table(tabla_pagos)
    # Aplicar bandas amarillas a filas pares de la tabla de pagos
    for row_idx in range(start_pagos + 1, end_pagos + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(
                    start_color=COLORES_INBLOQUET["amarillo"],
                    end_color=COLORES_INBLOQUET["amarillo"],
                    fill_type="solid"
                )

    # --- Tabla de Clases ---
    clases = alumno.get("clases", [])
    df_clases = pd.DataFrame(clases, columns=["fecha", "estado", "comentario"])
    if df_clases.empty:
        df_clases = pd.DataFrame([["", "", ""]], columns=["fecha", "estado", "comentario"])

    ws.append([])
    titulo_clases_row = ws.max_row + 1
    ws.cell(row=titulo_clases_row, column=1, value="Seguimiento de Clases").font = Font(bold=True, size=12, color=COLORES_INBLOQUET["azul_principal"])
    start_clases = ws.max_row + 1
    ws.append(["Fecha", "Estado", "Comentario"])
    for r in df_clases.itertuples(index=False):
        ws.append(list(r))
    end_clases = ws.max_row
    tabla_clases = Table(
        displayName=f"TablaClases_{alumno.get('Matricula','')}",
        ref=f"A{start_clases}:C{end_clases}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws.add_table(tabla_clases)
    # Aplicar bandas amarillas a filas pares de la tabla de clases
    for row_idx in range(start_clases + 1, end_clases + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(
                    start_color=COLORES_INBLOQUET["amarillo"],
                    end_color=COLORES_INBLOQUET["amarillo"],
                    fill_type="solid"
                )

    # --- Ajustar ancho de columnas (saltando celdas fusionadas) ---
    for col in ws.columns:
        valid_cells = [cell for cell in col if hasattr(cell, "column_letter") and cell.column_letter]
        if valid_cells:
            col_letter = valid_cells[0].column_letter
            max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col if hasattr(cell, "column_letter"))
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)


def exportar_todos_los_alumnos_inbloquet(alumnos: list, filename: str):
    """
    Exporta TODOS los alumnos, cada uno en su hoja, con el mismo formato 
    (encabezado, tres tablas para Datos, Pagos y Clases con bandas) usando TableStyleMedium9.
    Se corrigen los nombres de las tablas eliminando espacios.
    """
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    except locale.Error:
        locale.setlocale(locale.LC_TIME, "Spanish_Spain.1252")
    
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    for alumno in alumnos:
        hoja_titulo = f"{alumno.get('Nombre','Alumno')[:20]}"  # Limitar a 31 caracteres
        # Eliminar espacios para el nombre de la hoja si es necesario:
        ws = wb.create_sheet(title=hoja_titulo.replace(" ", "_"))
        
        # Encabezado (mismos merges y formato)
        ws.insert_rows(1, 3)
        ws.merge_cells("A1:B3")
        ws.merge_cells("D1:I1")
        ws.merge_cells("D2:I2")
        ws.merge_cells("D3:I3")
        ws.merge_cells("A4:K4")
        
        try:
            img = Image("InbloquetD.png")
            img.width = 120
            img.height = 120
            ws.add_image(img, "A1")
            row_height = img.height / 4
        except FileNotFoundError:
            ws["A1"] = "LOGO INBLOQUET"
            ws["A1"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["azul_principal"])
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            row_height = 20
        
        for i in range(1, 4):
            ws.row_dimensions[i].height = row_height

        ws["D1"] = "Ficha Informativa del Alumno"
        ws["D1"].font = Font(bold=True, size=20, color=COLORES_INBLOQUET["azul_principal"])
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["D2"] = f"ALUMNO: {alumno.get('Nombre','')}  -  Matrícula: {alumno.get('Matricula','')}"
        ws["D2"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["azul_secundario"])
        ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

        ws["D3"] = "¡Intenta, Explora y Conquista!"
        ws["D3"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["amarillo"])
        ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

        # --- Tabla de Datos ---
        campos = [
            ("Grado", alumno.get("Grado","")),
            ("Grado Inbloquet", alumno.get("Grado Inbloquet","")),
            ("Curso", alumno.get("Tipo de Curso","")),
            ("Fecha Inscripción", alumno.get("Fecha Inscripción","")),
            ("Fecha Inicio Clases", alumno.get("Fecha Inicio Clases","")),
            ("Estado Inscripción", alumno.get("Inscripción","")),
            ("Escuela Procedencia", alumno.get("Escuela de provinencia","")),
            ("Alergias", alumno.get("Alergias","Ninguna")),
            ("Observaciones", alumno.get("Observaciones","Ninguna")),
            ("Contacto Familiar", alumno.get("Nombre Completo del Familiar","")),
            ("Tel. Familiar", alumno.get("Número de teléfono del familiar","")),
            ("Contacto Emergencia", alumno.get("Nombre completo de contacto de emergencia","")),
            ("Tel. Emergencia", alumno.get("Número de teléfono de contacto de emergencia",""))
        ]
        start_data = ws.max_row + 1
        ws.append(["Campo", "Valor"])
        for campo in campos:
            ws.append(list(campo))
        end_data = ws.max_row
        tabla_datos = Table(
            displayName=f"TablaDatos_{hoja_titulo.replace(' ', '_')}",
            ref=f"A{start_data}:B{end_data}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        )
        ws.add_table(tabla_datos)
        for row_idx in range(start_data + 1, end_data + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill(
                        start_color=COLORES_INBLOQUET["amarillo"],
                        end_color=COLORES_INBLOQUET["amarillo"],
                        fill_type="solid"
                    )
        
        # --- Tabla de Pagos ---
        pagos = alumno.get("pagos", [])
        df_pagos = pd.DataFrame(pagos, columns=["fecha", "monto", "metodo"])
        if df_pagos.empty:
            df_pagos = pd.DataFrame([["", "", ""]], columns=["fecha", "monto", "metodo"])
        ws.append([])
        titulo_pagos_row = ws.max_row + 1
        ws.cell(row=titulo_pagos_row, column=1, value="Historial de Pagos").font = Font(bold=True, size=12, color=COLORES_INBLOQUET["azul_principal"])
        start_pagos = ws.max_row + 1
        ws.append(["Fecha", "Monto", "Método"])
        for r in df_pagos.itertuples(index=False):
            ws.append(list(r))
        end_pagos = ws.max_row
        tabla_pagos = Table(
            displayName=f"TablaPagos_{hoja_titulo.replace(' ', '_')}",
            ref=f"A{start_pagos}:C{end_pagos}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        )
        ws.add_table(tabla_pagos)
        for row_idx in range(start_pagos + 1, end_pagos + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill(
                        start_color=COLORES_INBLOQUET["amarillo"],
                        end_color=COLORES_INBLOQUET["amarillo"],
                        fill_type="solid"
                    )
        
        # --- Tabla de Clases ---
        clases = alumno.get("clases", [])
        df_clases = pd.DataFrame(clases, columns=["fecha", "estado", "comentario"])
        if df_clases.empty:
            df_clases = pd.DataFrame([["", "", ""]], columns=["fecha", "estado", "comentario"])
        ws.append([])
        titulo_clases_row = ws.max_row + 1
        ws.cell(row=titulo_clases_row, column=1, value="Seguimiento de Clases").font = Font(bold=True, size=12, color=COLORES_INBLOQUET["azul_principal"])
        start_clases = ws.max_row + 1
        ws.append(["Fecha", "Estado", "Comentario"])
        for r in df_clases.itertuples(index=False):
            ws.append(list(r))
        end_clases = ws.max_row
        tabla_clases = Table(
            displayName=f"TablaClases_{hoja_titulo.replace(' ', '_')}",
            ref=f"A{start_clases}:C{end_clases}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        )
        ws.add_table(tabla_clases)
        for row_idx in range(start_clases + 1, end_clases + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill(
                        start_color=COLORES_INBLOQUET["amarillo"],
                        end_color=COLORES_INBLOQUET["amarillo"],
                        fill_type="solid"
                    )
        
        # --- Ajuste de ancho de columnas (omitiendo celdas fusionadas) ---
        for col in ws.columns:
            valid_cells = [cell for cell in col if hasattr(cell, "column_letter") and cell.column_letter]
            if valid_cells:
                col_letter = valid_cells[0].column_letter
                max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col if hasattr(cell, "column_letter"))
                ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

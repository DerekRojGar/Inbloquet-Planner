# models/export_model.py
import locale
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image

# Paleta de colores corporativos
COLORES_INBLOQUET = {
    "azul_principal": "003891",
    "azul_secundario": "4BB1E0",
    "amarillo": "F6C500",
    "verde": "D1DD00"
}

def crear_excel_con_diseño(df, filename, semana, año, img_width=120, img_height=120):
    # Configurar localización
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    except locale.Error:
        locale.setlocale(locale.LC_TIME, "Spanish_Spain.1252")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Semana {semana}"
    
    # Configuración de encabezado
    ws.insert_rows(1, 3)
    ws.merge_cells("A1:B3")
    ws.merge_cells("D1:I1")
    ws.merge_cells("D2:I2")
    ws.merge_cells("D3:I3")
    ws.merge_cells("A4:K4")

    # Insertar solo logo INBLOQUET
    try:
        img = Image("InbloquetD.png")
        img.width = img_width
        img.height = img_height
        ws.add_image(img, "A1")
        row_height = img.height / 4
    except FileNotFoundError:
        ws["A1"] = "LOGO INBLOQUET"
        ws["A1"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["azul_principal"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row_height = 20

    ws.row_dimensions[1].height = row_height
    ws.row_dimensions[2].height = row_height
    ws.row_dimensions[3].height = row_height

    # Encabezado con estilos
    ws["D1"] = "Programación de Actividades Semanal"
    ws["D1"].font = Font(bold=True, size=20, color=COLORES_INBLOQUET["azul_principal"])
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D2"] = f"SEMANA {semana} - AÑO {año}"
    ws["D2"].font = Font(bold=True, size=16, color=COLORES_INBLOQUET["azul_secundario"])
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D3"] = "¡Intenta, Explora y Conquista!"
    ws["D3"].font = Font(bold=True, size=14, color=COLORES_INBLOQUET["amarillo"])
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # Formatear datos
    df["Fecha"] = df.apply(lambda row: format_fecha(row), axis=1)
    df = df.drop(columns=["Día", "Año"])
    df = df[["Semana", "Fecha", "Horario", "Alumnos", "Escuelas", "Grupos", "Maestro", "Tema", "Encargado", "Notas"]]

    # Añadir encabezados con estilo
    ws.append(df.columns.tolist())
    
    # Aplicar color a encabezados manualmente
    header_fill = PatternFill(
        start_color=COLORES_INBLOQUET["azul_principal"],
        end_color=COLORES_INBLOQUET["azul_principal"],
        fill_type="solid"
    )
    
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    # Añadir datos
    for row in df.to_dict('records'):
        ws.append(list(row.values()))
    
    # Crear tabla con estilo compatible
    tabla = Table(
        displayName="TablaActividades",
        ref=f"A5:J{ws.max_row}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws.add_table(tabla)

    # Aplicar color a filas alternas manualmente
    for row_idx in range(6, ws.max_row + 1):
        if row_idx % 2 == 0:  # Filas pares
            for cell in ws[row_idx]:
                cell.fill = PatternFill(
                    start_color=COLORES_INBLOQUET["amarillo"],
                    end_color=COLORES_INBLOQUET["amarillo"],
                    fill_type="solid"
                )

    # Ajustar columnas (solución al error de MergedCell)
    for col in ws.iter_cols(min_row=5, max_row=ws.max_row):  # Solo columnas con datos
        if col[0].column_letter:  # Verificar que no sea merged cell
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Ajustar texto en columna Tema
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(filename)

def format_fecha(row):
    try:
        return f"{row['Día']} {int(row['Fecha'][:2])} de " \
               f"{datetime.strptime(row['Fecha'], '%d/%m').strftime('%B').capitalize()} " \
               f"del {row['Año']}"
    except:
        return "Fecha no válida"


import pandas as pd
import streamlit as st
import base64
import os
import locale
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# ==================================================
# CONFIGURACIÓN INICIAL
# ==================================================
st.set_page_config(
    page_title="Planificador INBLOQUET",
    page_icon="📅",
    layout="wide"
)

# ==================================================
# ESTILOS CSS
# ==================================================
st.markdown("""
    <style>
    .logo-sidebar {
        max-width: 200px;
        margin: 20px auto;
        padding: 10px;
    }
    .frase-dia {
        background: #f8f9fa;
        padding: 20px;
        border-radius: 10px;
        border-left: 5px solid #003366;
        margin: 20px 0;
        color: #333;
    }
    .actividad-card {
        background: white;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin: 10px 0;
        color: #333; /* Color de texto oscuro para mejor legibilidad */
    }
    .actividad-card b {
        color: #003366; /* Color para los títulos */
    }
    .st-expander {
        background: white;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin: 10px 0;
    }
    .st-expander .stMarkdown {
        color: #333; /* Color de texto oscuro para mejor legibilidad */
    }
    </style>
    """, unsafe_allow_html=True)

# ==================================================
# FUNCIONES BASE
# ==================================================
def generar_semana(año, semana):
    fecha_inicio = datetime(año, 1, 4)
    fecha_inicio = fecha_inicio - timedelta(days=fecha_inicio.weekday())
    fecha_inicio += timedelta(weeks=semana-1)
    return [fecha_inicio + timedelta(days=i) for i in range(6)]  # Lunes a Sábado

DIAS_ESPAÑOL = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

ESCUELAS = ["INBLOQUET", "AEC", "RW Core", "RW Plus", "AB"]

GRUPOS_POR_ESCUELA = {
    "INBLOQUET": ["Dario", "Emi/Regi", "Roro Mine", "Santi", "Dani", "Romi", "Iker", "Hermanos"],
    "AEC": ["Taller 1", "Taller 2"],
    "RW Core": ["LOBOS", "RINOS", "PANDAS/BUFALOS", "PUMAS/DELFINES"],
    "RW Plus": ["S DUPLO", "S NORMAL", "M", "L"],
    "AB": ["PRESCO", "PA", "PB"]
}

if "escuelas_seleccionadas" not in st.session_state:
    st.session_state.escuelas_seleccionadas = []
if "grupos_disponibles" not in st.session_state:
    st.session_state.grupos_disponibles = []
if "buscar_grupo" not in st.session_state:
    st.session_state.buscar_grupos = False
# ==================================================
# CARGAR DATOS DESDE CSV
# ==================================================
def cargar_datos():
    if os.path.exists("actividades.csv"):
        try:
            df = pd.read_csv("actividades.csv")
            if df.empty:  # Verificar si el DataFrame está vacío
                return {}
            actividades = {}
            for _, row in df.iterrows():
                semana_key = f"{row['Año']}-S{row['Semana']}"
                fecha = row['Fecha']
                actividad = {
                    "Escuelas": row['Escuelas'],
                    "Horario": row ['Horario'],
                    "Grupos": row['Grupos'],
                    "Tema": row['Tema'],
                    "Encargado": row ['Encargado'],
                    "Maestro": row ['Maestro'],
                    "Alumnos": row['Alumnos'],
                    "Notas": row['Notas']
                }
                if semana_key not in actividades:
                    actividades[semana_key] = {
                        "fechas": [dia.strftime("%d/%m") for dia in generar_semana(int(row['Año']), int(row['Semana']))],
                        "actividades": {dia.strftime("%d/%m"): [] for dia in generar_semana(int(row['Año']), int(row['Semana']))}
                    }
                if fecha not in actividades[semana_key]["actividades"]:
                    actividades[semana_key]["actividades"][fecha] = []
                actividades[semana_key]["actividades"][fecha].append(actividad)
            return actividades
        except pd.errors.EmptyDataError:  # Manejar archivos vacíos
            return {}
    return {}

# ==================================================
# GUARDAR DATOS EN CSV
# ==================================================
def guardar_datos(actividades):
    all_data = []
    for semana_key, semana_data in actividades.items():
        año, semana = semana_key.split("-S")
        año = int(año)
        semana = int(semana)
        for fecha, actividades_dia in semana_data["actividades"].items():
            for act in actividades_dia:
                registro = {
                    "Semana": semana,
                    "Año": año,
                    "Fecha": fecha,
                    **act
                }
                all_data.append(registro)
    df = pd.DataFrame(all_data)
    df.to_csv("actividades.csv", index=False, encoding='utf-8-sig')

# ==================================================
# INICIALIZACIÓN DE DATOS
# ==================================================
if 'actividades' not in st.session_state:
    st.session_state.actividades = cargar_datos()

# ==================================================
# BARRA LATERAL (LOGO Y CONFIGURACIÓN)
# ==================================================
with st.sidebar:
    # Logo (reemplazar 'logo.png' con tu archivo)
    st.image("Inbloquet.png",  # 🖼️ Coloca tu archivo aquí
            use_container_width=True,  # Cambiado a use_container_width
            output_format='PNG',
            caption="",
            width=150,
            clamp=False)
    
    # Configuración de semana
    st.header("⚙️ Configuración")
    semana = st.number_input("Número de Semana", 1, 52, 9)
    año = st.number_input("Año", 2024, 2030, 2025)
    
    # Editor de frase
    st.markdown("---")
    frase_actual = st.session_state.get('frase_global', '')
    nueva_frase = st.text_area("✍️ Frase inspiradora del día:", value=frase_actual)
    if st.button("💾 Guardar Frase"):
        st.session_state.frase_global = nueva_frase
        st.rerun()

# ==================================================
# CONTENIDO PRINCIPAL
# ==================================================
st.title("📅 Planificación Semanal INBLOQUET")
st.markdown("---")

# Frase del día debajo del título
if 'frase_global' in st.session_state and st.session_state.frase_global:
    st.markdown(f'<div class="frase-dia">📌 {st.session_state.frase_global}</div>', 
               unsafe_allow_html=True)

# ==================================================
# VISTA DE CALENDARIO
# ==================================================
col1, col2, col3 = st.columns([6, 1, 1])  # Espacio para título y botones
with col1:
    st.subheader("🗓️ Vista Semanal")
with col2:
    if st.button("📂 Expandir Todo"):
        st.session_state.expanded_state = True
with col3:
    if st.button("📂 Contraer Todo"):
        st.session_state.expanded_state = False

# Inicializar el estado de expansión si no existe
if "expanded_state" not in st.session_state:
    st.session_state.expanded_state = False  # Por defecto, actividades cerradas

semana_key = f"{año}-S{semana}"
if semana_key not in st.session_state.actividades:
    st.session_state.actividades[semana_key] = {
        "fechas": [dia.strftime("%d/%m") for dia in generar_semana(año, semana)],
        "actividades": {dia.strftime("%d/%m"): [] for dia in generar_semana(año, semana)}
    }

cols_calendario = st.columns(6)
for i, col in enumerate(cols_calendario):
    with col:
        dia_str = st.session_state.actividades[semana_key]["fechas"][i]
        actividades = st.session_state.actividades[semana_key]["actividades"][dia_str]

        st.markdown(f"**{DIAS_ESPAÑOL[i]}**")
        st.caption(dia_str)

        if actividades:
            for idx, act in enumerate(actividades):
                # Separar escuelas y grupos
                escuelas_lista = act["Escuelas"].split(", ")
                grupos_lista = act["Grupos"].split(", ") if act["Grupos"] != "-" else ["Sin grupo"]

                for escuela, grupo in zip(escuelas_lista, grupos_lista):
                    # Formatear con colores para que solo aparezca en la parte superior
                    escuela_grupo_html = f"""
                    <span style="font-weight:bold; color:#003366;">{escuela}</span>: 
                    <span style="color:#FF5733; font-weight:bold;">{grupo}</span>
                    """

                    # Mostrar solo en el título del expander
                    with st.expander(label=f"{escuela}: {grupo}", expanded=st.session_state.expanded_state):
                        st.markdown(f"""
                        - **Horario:** {act['Horario']}
                        - **Alumnos:** {act['Alumnos']}
                        - **Grupos:** {act['Grupos']} 
                        - **Maestro:** {act['Maestro']}
                        - **Tema:** {act['Tema']}
                        - **Encargado:** {act['Encargado']}
                        """)

                        # Evitar claves duplicadas agregando escuela y grupo
                        if st.button(f"Detalles", key=f"btn_expand_{dia_str}_{idx}_{escuela}_{grupo}"):
                            st.session_state.selected_activity = {
                                "semana_key": semana_key,
                                "dia_str": dia_str,
                                "index": idx,
                                "datos": act
                            }
                            st.rerun()
        else:
            st.info("Sin actividades")


# ==================================================
# DETALLE DE ACTIVIDAD (DESPLAZAMIENTO AUTOMÁTICO)
# ==================================================
if "selected_activity" in st.session_state:
    selected = st.session_state.selected_activity
    act = selected["datos"]
    
    col1, col2 = st.columns([2, 7])  # Título a la izquierda, botón a la derecha
    with col1:
        st.subheader("📋 Detalle de Actividad")
    with col2:
        if st.button("❌ Ocultar"):
            del st.session_state.selected_activity
            st.rerun()

    st.markdown(f"""
    - **Horario:** {act['Horario']}
    - **Alumnos:** {act['Alumnos']}
    - **Escuelas:** {act['Escuelas']}
    - **Grupos:** {act['Grupos']}
    - **Maestro:** {act['Maestro']}
    - **Tema:** {act['Tema']}
    - **Encargado:** {act['Encargado']}
    - **Notas:** {act['Notas']}
    """)
    
    col1, col2 = st.columns([1, 6])
    
    with col1:
        if st.button("✏️ Editar"):
            st.session_state.editando = selected
            st.rerun()
    
    with col2:
        if st.button("🗑️ Eliminar"):
            semana_key = selected["semana_key"]
            dia_str = selected["dia_str"]
            index = selected["index"]

            # Eliminar la actividad de la estructura de datos
            del st.session_state.actividades[semana_key]["actividades"][dia_str][index]

            # ⚠️ Aquí aseguramos que el día NO SE ELIMINE aunque quede vacío
            # Simplemente dejamos la lista vacía para que se muestre el día sin actividades
            if not st.session_state.actividades[semana_key]["actividades"][dia_str]:
                st.session_state.actividades[semana_key]["actividades"][dia_str] = []

            # Guardar cambios en el archivo CSV eliminando la actividad
            guardar_datos(st.session_state.actividades)

            # Mensaje de éxito y recarga de la página
            st.success("✅ Actividad eliminada correctamente.")
            st.session_state.pop("selected_activity", None)  # Evitar acceso a actividad eliminada
            st.rerun()

# ==================================================
# MÓDULO DE EDICIÓN
# ==================================================
if 'editando' in st.session_state:
    st.markdown("---")
    st.subheader("✏️ Editor de Actividad")

    edit = st.session_state.editando

    with st.form(key='form_edicion'):
        cols = st.columns(2)

        with cols[0]:
            horario_edit = st.text_input(
                "Horario",
                value=edit["datos"]["Horario"] if edit["datos"]["Horario"] != "-" else ""
            )

            alumnos_edit = st.text_input(
                "Alumnos",
                value=edit["datos"]["Alumnos"] if edit["datos"]["Alumnos"] != "-" else ""
            )

            # MultiSelect de Escuelas con validación de opciones
            escuelas_disponibles = set(ESCUELAS)
            escuelas_existentes = [e for e in edit["datos"]["Escuelas"].split(", ") if e in escuelas_disponibles]

            escuelas_edit = st.multiselect(
                "Escuelas",
                list(ESCUELAS),
                default=escuelas_existentes
            )

            # Botón "Buscar Grupos" dentro del formulario
            if st.form_submit_button("🔍 Buscar grupos"):
                if escuelas_edit != st.session_state.get("escuelas_editando", []):
                    st.session_state.grupos_edit_disponibles = []
                    st.session_state.escuelas_editando = escuelas_edit

                # Filtra los grupos según las escuelas seleccionadas
                grupos_disponibles = set()
                for escuela in escuelas_edit:
                    grupos_disponibles.update(GRUPOS_POR_ESCUELA.get(escuela, []))

                # Guarda los grupos en session_state para persistencia
                st.session_state.grupos_edit_disponibles = sorted(grupos_disponibles)

                # Si no hay grupos, mostrar advertencia
                if not st.session_state.grupos_edit_disponibles:
                    st.warning("⚠️ No hay grupos disponibles para las escuelas seleccionadas.")

                st.rerun()  # Recarga para actualizar la lista de grupos

            # MultiSelect de Grupos basado en los datos guardados
            grupos_existentes = [g for g in edit["datos"]["Grupos"].split(", ") if g in st.session_state.get("grupos_edit_disponibles", [])]

            grupos_edit = st.multiselect(
                "Grupos",
                st.session_state.get("grupos_edit_disponibles", []),
                default=grupos_existentes
            )

        with cols[1]:
            maestro_edit = st.text_input(
                "Maestro",
                value=edit["datos"]["Maestro"] if edit["datos"]["Maestro"] != "-" else ""
            )

            tema_edit = st.text_area(
                "Tema",
                value=edit["datos"]["Tema"] if edit["datos"]["Tema"] != "-" else ""
            )

            encargado_edit = st.text_input(
                "Encargado",
                value=edit["datos"]["Encargado"] if edit["datos"]["Encargado"] != "-" else ""
            )

            notas_edit = st.text_area(
                "Notas",
                value=edit["datos"]["Notas"] if edit["datos"]["Notas"] != "-" else ""
            )

        # Botones dentro del formulario
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.form_submit_button("💾 Guardar Cambios"):
                if escuelas_edit:
                    nueva_actividad = {
                        "Horario": horario_edit.strip() if horario_edit else "-",
                        "Alumnos": alumnos_edit.strip(),
                        "Escuelas": ", ".join(escuelas_edit),
                        "Grupos": ", ".join(grupos_edit) if grupos_edit else "-",
                        "Maestro": maestro_edit.strip() if maestro_edit else "-",
                        "Tema": tema_edit.strip(),
                        "Encargado": encargado_edit.strip() if encargado_edit else "-",
                        "Notas": notas_edit.strip()
                    }
                    st.session_state.actividades[edit["semana_key"]]["actividades"][edit["dia_str"]][edit["index"]] = nueva_actividad
                    guardar_datos(st.session_state.actividades)
                    del st.session_state.editando
                    st.success("✅ Cambios guardados!")
                    st.rerun()
                else:
                    st.error("❌ Debes seleccionar al menos una escuela")

        with col2:
            if st.form_submit_button("❌ Cancelar Edición"):
                del st.session_state.editando
                st.rerun()

# ==================================================
# FORMULARIO DE NUEVA ACTIVIDAD
# ==================================================
st.markdown("---")
st.subheader("📝 Nueva Actividad")

with st.form(key='form_nueva_actividad'):
    cols = st.columns(2)

    with cols[0]:
        dia_seleccionado = st.selectbox("Selecciona el día:", DIAS_ESPAÑOL)

        horario = st.text_input(
            "Horario de la actividad:",
            placeholder="Ej: 8:00 a.m. - 2:00 p.m."
        )

        alumnos = st.text_input(
            "Alumnos:",
            placeholder="Ej: Dereck, Alberto, Emma"
        )

        # MultiSelect de Escuelas con búsqueda dinámica
        escuelas = st.multiselect(
            "Selecciona escuelas:",
            list(ESCUELAS),
            key='escuelas'
        )

        # Botón para buscar grupos según la selección de escuelas
        if st.form_submit_button("🔍 Buscar grupos"):
            if escuelas != st.session_state.get("escuelas_seleccionadas", []):
                st.session_state.grupos_disponibles = []
                st.session_state.escuelas_seleccionadas = escuelas

            # Filtra los grupos según las escuelas seleccionadas
            grupos_disponibles = set()
            for escuela in escuelas:
                grupos_disponibles.update(GRUPOS_POR_ESCUELA.get(escuela, []))

            # Guarda los grupos en session_state para persistencia
            st.session_state.grupos_disponibles = sorted(grupos_disponibles)

            # Si no hay grupos, mostrar advertencia
            if not st.session_state.grupos_disponibles:
                st.warning("⚠️ No hay grupos disponibles para las escuelas seleccionadas.")

            st.rerun()  # Recarga para actualizar la lista de grupos

        # MultiSelect de Grupos basado en la búsqueda
        grupos = st.multiselect(
            "Selecciona grupos:",
            st.session_state.get("grupos_disponibles", []),
            key='grupos'
        )

    with cols[1]:
        encargado = st.text_input(
            "Encargado",
            placeholder="Ej: Ivan, Gus, Caleb"
        )

        tema = st.text_area(
            "Descripción detallada:",
            placeholder="Ej: Taller de programación con robots",
            height=100
        )

        maestro = st.text_input(
            "Maestro",
            placeholder="Ej: Joss, Angie, Kevin"
        )

        notas = st.text_area(
            "Información adicional:",
            placeholder="Ej: Materiales especiales requeridos"
        )

    # Botón de envío dentro del formulario
    if st.form_submit_button("💾 Guardar Actividad"):
        if escuelas:
            nueva_actividad = {
                "Horario": horario if horario else "-",
                "Alumnos": alumnos if alumnos else "-",
                "Escuelas": ", ".join(escuelas),
                "Grupos": ", ".join(grupos) if grupos else "-",
                "Maestro": maestro if maestro else "-",
                "Tema": tema if tema else "-",
                "Encargado": encargado if encargado else "-",
                "Notas": notas if notas else "-"
            }
            semana_key = f"{año}-S{semana}"
            if semana_key not in st.session_state.actividades:
                st.session_state.actividades[semana_key] = {
                    "fechas": [dia.strftime("%d/%m") for dia in generar_semana(año, semana)],
                    "actividades": {dia.strftime("%d/%m"): [] for dia in generar_semana(año, semana)}
                }
            fecha_key = st.session_state.actividades[semana_key]["fechas"][DIAS_ESPAÑOL.index(dia_seleccionado)]
            st.session_state.actividades[semana_key]["actividades"][fecha_key].append(nueva_actividad)
            guardar_datos(st.session_state.actividades)
            st.success("✅ Actividad registrada exitosamente!")
            st.rerun()
        else:
            st.error("❌ Debes seleccionar al menos una escuela")

# ==================================================
# EXPORTACIÓN DE DATOS
# ==================================================
def crear_excel_con_diseño(df, filename, semana, año, img_width=120, img_height=120):
    # Configurar el idioma a español para los nombres de los meses
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")  # Linux/Mac
    except locale.Error:
        locale.setlocale(locale.LC_TIME, "Spanish_Spain.1252")  # Windows

    wb = Workbook()
    ws = wb.active
    ws.title = f"Semana {semana}"
    
    ws.insert_rows(1, 3)  # Insertamos 3 filas para el encabezado

    # COMBINAR CELDAS ===
    ws.merge_cells("A1:B3")  # Logo de la empresa
    ws.merge_cells("D1:I1")  # Información de la semana
    ws.merge_cells("D2:I2")
    ws.merge_cells("D3:I3")
    ws.merge_cells("J1:J3")
    #ws.merge_cells("A3:C3")
    ws.merge_cells("A4:K4")

    # INSERTAR LOGOS Y AJUSTAR TAMAÑO ===
    max_img_height = 0  # Variable para rastrear la imagen más grande en altura

    try:
        img1 = Image("InbloquetD.png")  # Cargar imagen del archivo
        img1.width = 150  # Tamaño personalizado
        img1.height = 140
        ws.add_image(img1, "A1")
        max_img_height = max(max_img_height, img1.height)
    except FileNotFoundError:
        ws["A1"] = "LOGO NO ENCONTRADO"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    try:
        img2 = Image("rocket.png")  # Cargar imagen del archivo
        img2.width = 90  # Tamaño personalizado
        img2.height = 140
        ws.add_image(img2, "J1")
        max_img_height = max(max_img_height, img2.height)
    except FileNotFoundError:
        ws["K1"] = "IMAGEN NO ENCONTRADA"
        ws["K1"].font = Font(bold=True, size=14)
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar la altura de las filas combinadas según la imagen más alta
    row_height = max_img_height / 4  # Ajuste proporcional
    ws.row_dimensions[1].height = row_height
    ws.row_dimensions[2].height = row_height
    ws.row_dimensions[3].height = row_height

    # INSERTAR ENCABEZADO PRINCIPAL ===
    ws["D1"] = "Programación de Actividades Semanal"
    ws["D1"].font = Font(bold=True, size=20)
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D2"] = f"SEMANA {semana} - AÑO {año}"
    ws["D2"].font = Font(bold=True, size=16)
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D3"] = "¡Intenta, Explora y Conquista!"
    ws["D3"].font = Font(bold=True, size=14)
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # REFORMATEAR COLUMNA FECHA ===
    df["Fecha"] = df.apply(lambda row: f"{row['Día']} {int(row['Fecha'][:2])} de {datetime.strptime(row['Fecha'], '%d/%m').strftime('%B')} del {row['Año']}", axis=1)
    
    # Convertir el mes a minúsculas y capitalizar la primera letra
    df["Fecha"] = df["Fecha"].apply(lambda x: x.replace(x.split()[3], x.split()[3].capitalize()))

    # Eliminar las columnas antiguas
    df = df.drop(columns=["Día", "Año"])

    # Reorganizar las columnas para que Fecha esté al inicio
    column_order = ["Semana", "Fecha", "Horario", "Alumnos", "Escuelas", "Grupos", "Maestro", "Tema", "Encargado", "Notas"]
    df = df[column_order]

    # INSERTAR ENCABEZADOS ===
    ws.append(df.columns.tolist())

    # INSERTAR DATOS ===
    for item in df.to_dict('records'):
        ws.append(list(item.values()))

    # FORMATEAR TABLA ===
    tabla_ref = f"A5:J{ws.max_row}"  # Tabla desde la fila 5 hasta la última con datos
    tabla = Table(displayName="TablaActividades", ref=tabla_ref)

    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,  
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    # AJUSTE AUTOMÁTICO DE COLUMNAS ===
    for col in ws.iter_cols(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        col_letter = col[0].column_letter
        max_length = max((len(str(celda.value)) if celda.value else 0) for celda in col) + 2
        ws.column_dimensions[col_letter].width = max_length

    # AJUSTAR TEXTO EN LA COLUMNA "TEMA" ===
    for celda in ws["H"]:  # Columna "Tema"
        celda.alignment = Alignment(wrap_text=True)

    # GUARDAR ARCHIVO EXCEL ===
    wb.save(filename)

# ==============================
# SECCIÓN DE EXPORTACIÓN
# ==============================

st.markdown("---")
if st.button("📤 Exportar a Excel y CSV"):
    all_data = []
    semana_data = st.session_state.actividades[semana_key]
    
    for fecha in semana_data["fechas"]:
        dia_nombre = DIAS_ESPAÑOL[semana_data["fechas"].index(fecha)]
        for act in semana_data["actividades"][fecha]:
            registro = {
                "Semana": semana,
                "Año": año,
                "Fecha": fecha,
                "Día": dia_nombre,
                **act
            }
            all_data.append(registro)
    
    df = pd.DataFrame(all_data)
    
    # Nombres de archivo
    excel_file = f"Planificacion_S{semana}_{año}.xlsx"
    csv_file = f"Planificacion_S{semana}_{año}.csv"
    
    # Guardar
    crear_excel_con_diseño(df, excel_file, semana, año, img_width=100, img_height=100)  # Aquí puedes cambiar el tamaño
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    
    # Descargas
    st.markdown("### 📥 Descargas") 
    st.markdown(f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64.b64encode(open(excel_file, "rb").read()).decode()}" download="{excel_file}">Descargar Excel</a>', unsafe_allow_html=True)
    st.markdown(f'<a href="data:text/csv;base64,{base64.b64encode(open(csv_file, "rb").read()).decode()}" download="{csv_file}">Descargar CSV</a>', unsafe_allow_html=True)
    st.success("✅ Exportación completada")

# ==================================================
# EXPORTACIÓN DE DATOS CON DISEÑO PARA TODAS LAS SEMANAS CON BARRA DE PROGRESO Y ORDENAMIENTO
# ==================================================
if st.button("📤 Exportar Semanas Totales"):
    wb = Workbook()
    # Eliminar la hoja por defecto
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Ordenar las semanas según año y número de semana
    semanas_ordenadas = sorted(
        st.session_state.actividades.keys(),
        key=lambda k: (int(k.split("-S")[0]), int(k.split("-S")[1]))
    )
    
    total_semanas = len(semanas_ordenadas)
    progreso = st.progress(0)
    
    for idx, semana_key in enumerate(semanas_ordenadas):
        # Extraer año y semana desde la clave (formato "AÑO-SSemana")
        año_str, semana_str = semana_key.split("-S")
        año = int(año_str)
        semana = int(semana_str)
        
        semana_data = st.session_state.actividades[semana_key]
        # Preparar los registros para la semana
        all_data = []
        for fecha in semana_data["fechas"]:
            dia_nombre = DIAS_ESPAÑOL[semana_data["fechas"].index(fecha)]
            for act in semana_data["actividades"][fecha]:
                registro = {
                    "Semana": semana,
                    "Año": año,
                    "Fecha": fecha,
                    "Día": dia_nombre,
                    **act
                }
                all_data.append(registro)
        
        # Si existen actividades, crear la hoja con diseño
        if all_data:
            df = pd.DataFrame(all_data)
            # Crear hoja con nombre identificativo, por ejemplo: "S9_2025"
            sheet_name = f"S{semana}_{año}"
            ws = wb.create_sheet(title=sheet_name)
            
            # ===== Encabezado y estilos =====
            ws.insert_rows(1, 3)  # Reservar 3 filas para el encabezado
            
            # Combinar celdas para logos e información
            ws.merge_cells("A1:B3")  # Logo de la empresa
            ws.merge_cells("D1:I1")  # Información de la semana
            ws.merge_cells("D2:I2")
            ws.merge_cells("D3:I3")
            ws.merge_cells("J1:J3")
            ws.merge_cells("A4:K4")
            
            # Insertar logos y ajustar tamaño
            max_img_height = 0
            try:
                img1 = Image("InbloquetD.png")
                img1.width = 150
                img1.height = 140
                ws.add_image(img1, "A1")
                max_img_height = max(max_img_height, img1.height)
            except FileNotFoundError:
                ws["A1"] = "LOGO NO ENCONTRADO"
                ws["A1"].font = Font(bold=True, size=14)
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            try:
                img2 = Image("rocket.png")
                img2.width = 90
                img2.height = 140
                ws.add_image(img2, "J1")
                max_img_height = max(max_img_height, img2.height)
            except FileNotFoundError:
                ws["K1"] = "IMAGEN NO ENCONTRADA"
                ws["K1"].font = Font(bold=True, size=14)
                ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar la altura de las filas según la imagen más alta
            row_height = max_img_height / 4
            ws.row_dimensions[1].height = row_height
            ws.row_dimensions[2].height = row_height
            ws.row_dimensions[3].height = row_height
            
            # Insertar textos del encabezado
            ws["D1"] = "Programación de Actividades Semanal"
            ws["D1"].font = Font(bold=True, size=20)
            ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["D2"] = f"SEMANA {semana} - AÑO {año}"
            ws["D2"].font = Font(bold=True, size=16)
            ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
            ws["D3"] = "¡Intenta, Explora y Conquista!"
            ws["D3"].font = Font(bold=True, size=14)
            ws["D3"].alignment = Alignment(horizontal="center", vertical="center")
            
            # ===== Procesar el DataFrame =====
            # Reformatear la columna Fecha (por ejemplo: "Lunes 01 de Enero del 2025")
            df["Fecha"] = df.apply(lambda row: f"{row['Día']} {int(row['Fecha'][:2])} de {datetime.strptime(row['Fecha'], '%d/%m').strftime('%B')} del {row['Año']}", axis=1)
            df["Fecha"] = df["Fecha"].apply(lambda x: x.replace(x.split()[3], x.split()[3].capitalize()))
            # Eliminar columnas que ya no se usarán
            df = df.drop(columns=["Día", "Año"])
            # Reorganizar las columnas para que "Fecha" quede al inicio
            column_order = ["Semana", "Fecha", "Horario", "Alumnos", "Escuelas", "Grupos", "Maestro", "Tema", "Encargado", "Notas"]
            df = df[column_order]
            
            # ===== Insertar datos en la hoja =====
            # Insertar encabezados de la tabla
            ws.append(df.columns.tolist())
            # Insertar cada registro
            for item in df.to_dict('records'):
                ws.append(list(item.values()))
            
            # ===== Aplicar formato a la tabla =====
            last_row = ws.max_row
            tabla_ref = f"A5:J{last_row}"
            tabla = Table(displayName=f"TablaActividades_{sheet_name}", ref=tabla_ref)
            estilo_tabla = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo_tabla
            ws.add_table(tabla)
            
            # Ajuste automático del ancho de columnas
            for col in ws.iter_cols(min_row=5, max_row=last_row, min_col=1, max_col=ws.max_column):
                col_letter = col[0].column_letter
                max_length = max((len(str(celda.value)) if celda.value else 0) for celda in col) + 2
                ws.column_dimensions[col_letter].width = max_length
            
            # Ajustar texto en la columna "Tema" (columna H)
            for celda in ws["H"]:
                celda.alignment = Alignment(wrap_text=True)
        
        # Actualizar la barra de progreso
        progreso.progress((idx + 1) / total_semanas)
    
    # Guardar el archivo Excel con diseño para todas las semanas
    excel_file = "Planificacion General de Semanas.xlsx"
    wb.save(excel_file)
    
    # Eliminar la barra de progreso (opcional)
    progreso.empty()
    
    # Generar enlace de descarga
    with open(excel_file, "rb") as f:
        excel_bytes = f.read()
    excel_base64 = base64.b64encode(excel_bytes).decode()
    st.markdown("### 📥 Descarga")
    st.markdown(
        f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{excel_file}">Descargar Excel</a>',
        unsafe_allow_html=True
    )
    st.success("✅ Exportación completada")



# ==================================================
# INSTRUCCIONES FINALES
# ==================================================



# Ejecutar con:
# python -m streamlit run planificador.py